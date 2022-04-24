from openpyxl import Workbook, load_workbook

hero = ['susuan','may','juliet','samson','moses','jesus','hippopotamus']
def get_value(letter,num):
    value = (str(letter) + str(num))
    cell = ws[value]
    print(cell)

def save1():
    wb.save("excel.xlsx")

def save2(name):
    WB.save(name)

wb = load_workbook("excel.xlsx")
ws = wb.active

# get value
print(ws['A1'].value)

wb.create_sheet("test")
wb.create_sheet("tes23")
wb.create_sheet("te23t")
wb.create_sheet("te1121t")
wb.create_sheet("te123")

print(wb.sheetnames)

WB = Workbook()
WS = WB.active

WS.title = "Students"


print(WB.sheetnames)
WS.append(hero)



save2("student_test_save.xlsx")









